# Bruno Iochins Grisci
# May 24th, 2020

try:
    import sys
except:
    print('Could not import sys')

if sys.version_info < (3, 0):
    sys.stdout.write("Sorry, requires Python 3.x, not Python 2.x\n")
    sys.exit(1)

########################################

try:
    import os
except:
    print('Could not import os')
try:
    import importlib
except:
    print('Could not import importlib')
try:
    import importlib.util
except:
    print('Could not import importlib.util')
try:
    from collections import namedtuple
except:
    print('Could not import namedtuple from collections')
try:
    import pickle
except:
    print('Could not import pickle')
try:
    import itertools
except:
    print('Could not import itertools')
try:
    import tempfile
except:
    print('Could not import tempfile')

try:
    import importlib.util as imp
except:
    print('Could not import importlib.util')
try:
    if imp.find_spec("cupy"):
        import cupy
        import cupy as np
except:
    print('Could not import cupy')
    
try:
    from tqdm import tqdm
except:
    print('Could not import tqdm')

########################################

try:
    import matplotlib
except:
    print('Could not import matplotlib')

matplotlib.use('Agg')

try:
    import matplotlib.pyplot as plt
except:
    print('Could not import matplotlib.pyplot')

########################################

try:
    import numpy as np
except:
    print('Could not import numpy')

try:
    import pandas as pd
except:
    print('Could not import pandas')

print('Pandas version must be >= 0.25.3, it is {}'.format(pd.__version__))

try:
    import seaborn as sn
except:
    print('Could not import seaborn')

########################################

try:
    import scipy.stats as stats
except:
    print('Could not import scipy.stats')
try:
    from scipy.stats import gmean
except:
    print('Could not import gmean from scipy.stats')
try:
    from scipy.stats.mstats import rankdata
except:
    print('Could not import rankdata from scipy.stats.mstats')

########################################

try:
    import sklearn
except:
    print('Could not import sklearn')    
try:
    from sklearn import metrics
except:
    print('Could not import metrics from sklearn')
try:
    from sklearn import datasets
except:
    print('Could not import datasets from sklearn')
try:
    from sklearn.preprocessing import StandardScaler
except:
    print('Could not import StandardScaler from sklearn.preprocessing')
try:
    from sklearn.decomposition import PCA
except:
    print('Could not import PCA from sklearn.decomposition')
try:
    from sklearn.manifold import TSNE
except:
    print('Could not import TSNE from sklearn.manifold')
try:
    from sklearn.utils import class_weight
except:
    print('Could not import class_weight from sklearn.utils')
try:
    from sklearn.model_selection import StratifiedKFold, KFold
except:
    print('Could not import StratifiedKFold, KFold from sklearn.model_selection')    

########################################

try:
    from openpyxl import Workbook
except:
    print('Could not import Workbook from openpyxl')    
try:
    from openpyxl.styles import PatternFill
except:
    print('Could not import PatternFill from openpyxl.styles')
try:
    from openpyxl.styles import Font, Color
except:
    print('Could not import Font, Color from openpyxl.styles')
try:
    from openpyxl.styles import Alignment
except:
    print('Could not import Alignment from openpyxl.styles')
try: 
    from openpyxl.utils import get_column_letter
except ImportError:
    from openpyxl.cell import get_column_letter
except:
    print('Could not import get_column_letter from openpyxl')

########################################

try:
    import tensorflow
except:
    print('Could not import tensorflow') 
try:    
    from tensorflow.keras import backend as K
except:
    print('Could not import backend from tensorflow.keras')     
try:
    from tensorflow.keras import activations
except:
    print('Could not import activations from tensorflow.keras') 
try:
    from tensorflow.keras.models import load_model
except:
    print('Could not import load_model from tensorflow.keras.models') 
try:
    from tensorflow.keras.models import Sequential
except:
    print('Could not import Sequential from tensorflow.keras.models') 
try:
    from tensorflow.keras.layers import Dense, Activation, Dropout
except:
    print('Could not import Dense, Activation, Dropout from tensorflow.keras.layers') 
try:
    from tensorflow.keras.constraints import Constraint
except:
    print('Could not import Constraint from tensorflow.keras.constraints') 
try:
    from tensorflow.keras.constraints import max_norm
except:
    print('Could not import max_norm from tensorflow.keras.constraints') 
try:
    from tensorflow.keras.constraints import non_neg
except:
    print('Could not import non_neg from tensorflow.keras.constraints') 
try:
    from tensorflow.keras.utils import plot_model, to_categorical
except:
    print('Could not import plot_model, to_categorical from tensorflow.keras.utils') 
try:
    from tensorflow.keras.optimizers import Adam
except:
    print('Could not import Adam from tensorflow.keras.optimizers') 

tensorflow.compat.v1.enable_eager_execution() 
########################################

try:
    import deeplift
except:
    print('Could not import deeplift')    
try:
    from deeplift.conversion import kerasapi_conversion as kc
except:
    print('Could not import kerasapi_conversion from deeplift.conversion')  

########################################

try:
    import model_io
except:
    print('Could not import model_io')
try:
    import RR_utils
except:
    print('Could not import RR_utils')
try:
    import plot_pca
except:
    print('Could not import plot_pca')
try:
    from modules import Sequential,Linear,Tanh,Rect,SoftMax,Convolution,Flatten,SumPool,MaxPool
except:
    print('Could not import Sequential,Linear,Tanh,Rect,SoftMax,Convolution,Flatten,SumPool,MaxPool from modules')

print('###############')
print('Finished checking dependencies. Please install any missing library!')