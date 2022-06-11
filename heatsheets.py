# Bruno Iochins Grisci
# May 24th, 2020

import os
import sys
import importlib
import numpy as np
import pandas as pd
import seaborn as sn
from scipy.stats.mstats import rankdata 
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font, Color
from openpyxl.styles import Alignment

try: 
    from openpyxl.utils import get_column_letter
except ImportError:
    from openpyxl.cell import get_column_letter

import RR_utils

config_file = sys.argv[1]
cfg = importlib.import_module(config_file.replace('/','.').replace('.py',''))

colhex = {
    'RED':     'BA0000',
    'BLUE':    '0000FF',
    'YELLOW':  'FFEE00',
    'GREEN':   '048200',    
    'ORANGE':  'FF6103',
    'BLACK':   '000000',
    'CYAN':    '00FFD4',    
    'SILVER':  'c0c0c0',
    'MAGENTA': '680082',
    'CREAM':   'FFFDD0',
    'DRKBRW':  '654321',
    'BEIGE':   'C2C237',
    'WHITE':   'FFFFFF',
}

if cfg.class_colors is None:
    CLASS_COLORS = [PatternFill(fgColor=colhex['CYAN'], fill_type = 'solid'), PatternFill(fgColor=colhex['ORANGE'], fill_type = 'solid'), PatternFill(fgColor=colhex['DRKBRW'], fill_type = 'solid'), PatternFill(fgColor=colhex['ORANGE'], fill_type = 'solid'), PatternFill(fgColor=colhex['CYAN'], fill_type = 'solid')]
else:
    CLASS_COLORS = []
    for c in cfg.class_colors:
        CLASS_COLORS.append(PatternFill(fgColor=colhex[c], fill_type = 'solid'))

USES_COLORS  = [PatternFill(fgColor=colhex['BLACK'], fill_type = 'solid'), PatternFill(fgColor=colhex['SILVER'], fill_type = 'solid')]
TRAINTEST = ['train', 'test']

SCORE_LABEL = 'score'

####################################################################################

def hex2rgb(color):
    return tuple(int(color[i:i+2], 16) for i in (0, 2, 4))

def rgb2hex(color):
    return '%02x%02x%02x' % color

def computeRGB(c1=(255,255,255), c2=(0,0,0), percent=1.0):
    r1 = c1[0]/255.0
    r2 = c2[0]/255.0
    g1 = c1[1]/255.0
    g2 = c2[1]/255.0
    b1 = c1[2]/255.0
    b2 = c2[2]/255.0        
    r = r1 + percent * (r2 - r1)
    g = g1 + percent * (g2 - g1)
    b = b1 + percent * (b2 - b1)
    return (int(r*255),int(g*255),int(b*255))

def paint(cell, hex0, alpha):
    alpha=float(alpha)
    rgb0 = hex2rgb(hex0)
    if np.isnan(alpha):
        rgb1 = computeRGB(c2=rgb0, percent=0.0)
    else:
        rgb1 = computeRGB(c2=rgb0, percent=alpha)
    hex1 = rgb2hex(rgb1)
    pf   = PatternFill(fgColor=hex1, fill_type = 'solid')
    cell.fill = pf

def paint_class(cell, index, condition, class_type, alpha=1.0):
    if index in condition:
        color_index = condition.index(index)%len(class_type)
        hex0 = class_type[color_index].fgColor.rgb[2:]
        paint(cell, hex0, alpha)

if __name__ == '__main__':

    out_fold, out_file = RR_utils.create_output_dir(cfg.dataset_file, cfg.output_folder)
    if not os.path.exists(out_fold+'relevance_eval/'):
        os.makedirs(out_fold+'relevance_eval/')
    df = pd.read_csv(cfg.dataset_file, delimiter=cfg.dataset_sep, header=0, index_col=cfg.row_index)
    df = RR_utils.check_dataframe(df, cfg.class_label, cfg.task)

    if cfg.task == 'classification':
        CLASS_LABELS = list(np.sort(df[cfg.class_label].astype(str).unique()))
    elif cfg.task == 'regression':
        CLASS_LABELS, _ = RR_utils.split_targets(df, df[cfg.class_label].astype(float).min(), df[cfg.class_label].astype(float).max(), cfg.target_split, cfg.class_label)
        CLASS_LABELS = list(CLASS_LABELS)
    else:
        raise Exception('Unknown task type: {}'.format(cfg.task))    
    n_classes  = len(CLASS_LABELS)
    n_features = len(df.columns.values) - 1

    max_print_feats = 1000
    if n_features > max_print_feats:
        print('\nWARNING: Over {} features, only the top {} features will be printed in the sheet (it does not affect the averages).\n'.format(max_print_feats, max_print_feats))

    if cfg.k > 1:
        usage = TRAINTEST
    else:
        usage = TRAINTEST[0:1]
    for fold in range(max(cfg.k, 1)):
        wb = Workbook()
        sheets = {TRAINTEST[0]: wb.active, TRAINTEST[1]: wb.create_sheet()}
        for use in usage:
            print('\n###### {}-FOLD: {}\n'.format(fold+1, use))

            data_file = '{}_{}_{:04d}_{}_datasheet.csv'.format(out_file, fold+1, cfg.train_epochs, use)
            rele_file = '{}_{}_{:04d}_{}_relsheet.csv'.format(out_file, fold+1, cfg.train_epochs, use)

            dat = pd.read_csv(data_file, header=0, index_col=0, low_memory=False)
            rel = pd.read_csv(rele_file, header=0, index_col=0, low_memory=False)

            max_print_samples = 500
            if len(dat.columns) > max_print_samples+len(CLASS_LABELS)+1:
                print(dat.columns)
                print('\nWARNING: Over {} samples, only {} random samples will be printed in the sheet (it does not affect the averages).\n'.format(max_print_samples, max_print_samples))
                random_samples = dat.sample(max_print_samples, axis=1)
                random_samples = random_samples.sort_values(by=['usage', cfg.class_label, 'prediction', 'max_out'], axis=1, ascending=True, na_position='first')
                random_samples = list(random_samples.columns)
                use_cols = [SCORE_LABEL] + CLASS_LABELS + random_samples
                seen = set()
                use_cols = [x for x in use_cols if not (x in seen or seen.add(x))]
                dat = dat[use_cols]
                rel = rel[use_cols]
                print(dat.columns)
                print(len(dat.columns))

            print('\nData:\n')
            print(dat)
            print('\nRelevance:\n')
            print(rel)

            USAGE_row  = 2
            if cfg.task == 'classification':
                nc = n_classes
            elif cfg.task == 'regression':
                nc = 1
            PROBS_rows = list(range(USAGE_row+1, USAGE_row+1+nc))
            BRUTE_rows = list(range(max(PROBS_rows)+1, max(PROBS_rows)+1+nc))              
            MAX_row    = max(BRUTE_rows) + 1
            PRED_row   = MAX_row + 1
            CLASS_row  = PRED_row + 1
            INTRO_rows = list(range(USAGE_row, CLASS_row+1))
            FEAT_rows  = list(range(max(INTRO_rows)+1, len(dat.index)+2))

            AVER_col    = 2
            CLASS_cols  = list(range(AVER_col+1, AVER_col+1+n_classes))

            SAMPLE_cols = list(range(max(CLASS_cols)+1, len(dat.columns)+2))

            sheets[use].title = use

            # COLUMNS HEADER
            ir, ic = 1, AVER_col
            for col in dat.columns.values:
                c = sheets[use].cell(row=ir, column=ic, value=col)
                c.font = c.font.copy(bold=True)
                paint_class(c, ic, CLASS_cols, CLASS_COLORS)
                ic = ic+1

            # ROWS HEADER
            ir, ic = USAGE_row, 1
            for row in dat.index.values:
                c = sheets[use].cell(row=ir, column=ic, value=row)
                c.font = c.font.copy(bold=True)
                paint_class(c, ir, PROBS_rows, CLASS_COLORS)
                paint_class(c, ir, BRUTE_rows, CLASS_COLORS)
                ir = ir+1        

            sheets[use].freeze_panes = '{}{}'.format(get_column_letter(min(SAMPLE_cols)), min(FEAT_rows))

            # GLOBAL AVERAGE RANKING
            ir, ic = USAGE_row, AVER_col
            for av in dat[SCORE_LABEL]:
                c = sheets[use].cell(row=ir, column=ic, value=av)
                if cfg.rank == 'rank':
                    av_alpha = 1.0-(float(av)/n_features)
                elif cfg.rank == 'norm':
                    av_alpha = float(av)
                else:
                    raise Exception('Unknown rank method: {}'.format(cfg.rank))
                paint(c, PatternFill(fgColor=colhex['SILVER'], fill_type = 'solid').fgColor.rgb[2:], alpha=av_alpha)
                ir = ir+1
            
            # CLASS AVERAGE RANKING
            for ic in CLASS_cols:
                for ir in FEAT_rows:
                    avg_rank = dat.iloc[ir-2, ic-2]
                    c = sheets[use].cell(row=ir, column=ic, value=avg_rank)
                    if cfg.rank == 'rank':
                        avg_alpha = 1.0-(float(avg_rank)/n_features)
                    elif cfg.rank == 'norm':
                        avg_alpha = float(avg_rank)
                    else:
                        raise Exception('Unknown rank method: {}'.format(cfg.rank))                    
                    paint_class(c, ic, CLASS_cols, CLASS_COLORS, alpha=avg_alpha)

            # SAMPLES INFORMATION
            for ic in SAMPLE_cols:
                for ir in INTRO_rows:
                    info = dat.iloc[ir-2, ic-2]
                    if ir in PROBS_rows or ir in BRUTE_rows or ir == MAX_row:
                        if cfg.task == 'classification':
                            c = sheets[use].cell(row=ir, column=ic, value=float(info))
                        elif cfg.task == 'regression':
                            c = sheets[use].cell(row=ir, column=ic, value=info)
                    else:
                        c = sheets[use].cell(row=ir, column=ic, value=info)
                    if ir == USAGE_row:
                        paint(c, USES_COLORS[TRAINTEST.index(info)].fgColor.rgb[2:], alpha=1.0)
                    if cfg.task == 'classification':
                        paint_class(c, ir, PROBS_rows, CLASS_COLORS, alpha=info)
                        paint_class(c, ir, BRUTE_rows, CLASS_COLORS, alpha=dat.iloc[ir-2-n_classes, ic-2])
                    elif cfg.task == 'regression':
                        paint_class(c, ir, PROBS_rows, CLASS_COLORS)
                        paint_class(c, ir, BRUTE_rows, CLASS_COLORS)
                    if ir == MAX_row:
                        if cfg.task == 'classification':
                            paint(c, CLASS_COLORS[CLASS_LABELS.index(dat.iloc[PRED_row-2, ic-2])].fgColor.rgb[2:], alpha=1.0)
                        elif cfg.task == 'regression':
                            paint(c, CLASS_COLORS[0].fgColor.rgb[2:], alpha=1.0)
                    if ir == PRED_row:
                        if cfg.task == 'classification':
                            paint(c, CLASS_COLORS[CLASS_LABELS.index(info)].fgColor.rgb[2:], alpha=1.0)
                        elif cfg.task == 'regression':
                            paint(c, CLASS_COLORS[0].fgColor.rgb[2:], alpha=1.0)
                    if ir == CLASS_row:
                        if cfg.task == 'classification':
                            paint(c, CLASS_COLORS[CLASS_LABELS.index(info)].fgColor.rgb[2:], alpha=1.0)
                        elif cfg.task == 'regression':
                            paint(c, CLASS_COLORS[0].fgColor.rgb[2:], alpha=1.0)

            # DATA AND RELEVANCE
            for ic in SAMPLE_cols:
                max_relevance = rel.iloc[min(FEAT_rows)-2:, ic-2].astype(float).abs().max()
                for ir in FEAT_rows[0:min(max_print_feats, max(FEAT_rows))]:
                    try:
                        v = float(dat.iloc[ir-2, ic-2])
                    except:
                        v = dat.iloc[ir-2, ic-2]
                    r = float(rel.iloc[ir-2, ic-2])
                    c = sheets[use].cell(row=ir, column=ic, value=v)
                    if r > 0.0:
                        r_col = PatternFill(fgColor=colhex['RED'], fill_type = 'solid').fgColor.rgb[2:]
                    else:
                        r_col = PatternFill(fgColor=colhex['BLUE'], fill_type = 'solid').fgColor.rgb[2:]
                    if max_relevance != 0.0:
                        if abs(r)/max_relevance > 0.5:
                            c.font = Font(color=colhex['WHITE'])
                        paint(c, r_col, alpha=abs(r)/max_relevance)
                    else:
                        paint(c, r_col, alpha=0.0)

            df_selection = pd.DataFrame([], columns=[SCORE_LABEL] + CLASS_LABELS)
            for cl in [SCORE_LABEL] + CLASS_LABELS:
                rank = dat[cl].copy()
                if cfg.rank == 'rank':
                    rank = rank.sort_values(ascending=True, na_position='last')
                elif cfg.rank == 'norm':
                    rank = rank.sort_values(ascending=False, na_position='last')
                else:
                    raise Exception('Unknown rank method: {}'.format(cfg.rank))
                rank_id = list(rank.index.values)[:cfg.n_selection]
                df_selection[cl] = rank_id

            all_ids = []
            i = 0
            while (len(all_ids) < max(cfg.n_selection, 2*len(CLASS_LABELS))) and i < cfg.n_selection:
                for cl in CLASS_LABELS:
                    all_ids.append(df_selection[cl].values[i])
                    all_ids = list(set(all_ids))
                i = i+1

            all_ids = list(set(all_ids))

            print('\nTop {} features by class:\n'.format(cfg.n_selection))
            print(df_selection)
            df_selection.to_csv('{}{}_{:04d}_{}_selection.csv'.format(out_fold+'relevance_eval/', fold+1, cfg.train_epochs, use))

            class_venn = RR_utils.venn([list(df_selection[cl].values) for cl in [SCORE_LABEL]+CLASS_LABELS], [SCORE_LABEL]+CLASS_LABELS, cfg.n_selection)
            print('\nSet difference between classes (top {} features):\n'.format(cfg.n_selection))
            print(class_venn)
            class_venn.to_csv('{}{}_{:04d}_{}_class_venn.csv'.format(out_fold+'relevance_eval/', fold+1, cfg.train_epochs, use))

            cat_all_ids = []
            if cfg.agglutinate:
                for aid in all_ids:
                    if aid in list(df.columns.values):
                        cat_all_ids.append(aid)
                    else:
                        for col in list(df.columns.values):
                            if aid + '.' in col:
                                cat_all_ids.append(col)
            else:
                cat_all_ids = all_ids

            if cfg.task == 'classification':
                correlation = df[cat_all_ids].corr(method ='pearson')
            elif cfg.task == 'regression':
                df[cfg.class_label] = df[cfg.class_label].astype(float)
                c_sel = [cfg.class_label]+cat_all_ids
                correlation = df[c_sel].corr(method ='pearson')
            print('\nCorrelation: \n{}'.format(correlation))
            correlation.to_csv('{}{}_{:04d}_{}_corr.csv'.format(out_fold+'relevance_eval/', fold+1, cfg.train_epochs, use))

        wb.save(data_file.replace('.csv', '.xlsx'))